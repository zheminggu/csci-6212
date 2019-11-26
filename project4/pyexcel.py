from openpyxl import Workbook, load_workbook


class Excel_dealer(object):
    def __init__(self, filename):
        self.filename = filename
        try:
            self.workbook = load_workbook(filename=f"{self.filename}.xlsx")
        except EnvironmentError as e:
            print(f"on init {e}")
            self.workbook = Workbook()

        self.sheet = self.workbook.active

    def insert(self, column, row, information):
        coodinate = f"{column}{row}"
        self.sheet[coodinate] = f"{information}"

    def savefile(self):
        self.workbook.save(filename=f"{self.filename}.xlsx")


if __name__ == "__main__":
    excel_dealer = Excel_dealer("Helloworld")
    excel_dealer.insert("A", 1, "Hello")
    excel_dealer.insert("B", 1, "World")
    excel_dealer.savefile()
